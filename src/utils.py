import logging
import os
import shutil
import time

import pyautogui
from openpyxl import Workbook, load_workbook

from src.image_processing import get_screenshot


def clear_screenshots_directory():
    """
    Clears the screenshots directory by removing all files and recreating the directory.
    """

    directory = '../logs/screenshots'

    if os.path.exists(directory):
        shutil.rmtree(directory)
        os.makedirs(directory)
        with open(os.path.join(directory, '.gitkeep'), 'w'):
            pass

        logging.info(f"Directory '{directory}' has been cleared.")
    else:
        logging.info(f"Directory '{directory}' does not exist.")


# Define button coordinates as constants
ATTACK_BUTTON = (935, 1365)
FIND_TARGET_BUTTON = (950, 1270)
FIGHT_NOW_BUTTON = (1360, 865)
SEARCH_AGAIN_BUTTON = (1030, 1305)
TRAINING_CAMP_1_BUTTON = (1265, 505)
TRAINING_CAMP_2_BUTTON = (1390, 565)
CLOSE_TRAINING_VIEW_BUTTON = (1615, 1340)
SPEED_UP_X2_BUTTON = (1535, 160)
END_BATTLE_BUTTON = (1475, 200)
GO_HOME_BUTTON = (1250, 935)
OPEN_PLANETS_LIST_BUTTON = (940, 1320)
COLONY_11_BUTTON = (1190, 1240)
ADD_LOOTERS_TO_TRAINING_LIST_BUTTON = (1070, 1420)
CHOOSE_LOOTER_UNIT_WHEN_ATTACKING_BUTTON = (950, 1375)
CLOSE_NEWS_POPUP_BUTTON = (1583, 204)
CLOSE_DAILY_GIFT_POPUP_BUTTON = (1590, 530)


def click_and_wait(button, time_to_wait):
    """
    Clicks a specified button and waits for a given amount of time.

    Params:
        button (tuple): The (x, y) coordinates of the button to click.
        time_to_wait (float): The number of seconds to wait after clicking.
    """

    pyautogui.click(button)
    time.sleep(time_to_wait)


def get_initial_base():
    """
    Navigates to the initial base.
    """

    try:
        click_and_wait(OPEN_PLANETS_LIST_BUTTON, 2)
        click_and_wait(COLONY_11_BUTTON, 7)
    except Exception as e:
        raise e


def handle_error():
    """
    Handles errors by refreshing the game window and resetting to the initial base.
    """

    get_screenshot("Galaxy Life")  # take focus on window
    logging.warning("---------------------------------Handling error with F5 refresh---------------------------------")
    pyautogui.keyDown('F5')
    time.sleep(0.2)
    pyautogui.keyUp('F5')
    time.sleep(15)
    click_and_wait(CLOSE_NEWS_POPUP_BUTTON, 3)
    click_and_wait(CLOSE_DAILY_GIFT_POPUP_BUTTON, 2)

    get_initial_base()


class ExcelLogger:
    def __init__(self, init_time):
        self.init_time = init_time
        self.workbook = None
        self.sheet = None
        self.filename = None
        self.init_excel_logging()

    def init_excel_logging(self):
        """
        Initializes the Excel workbook for logging.
        """
        init_time_formatted = self.init_time.strftime("%Y%m%d_%H%M%S")
        self.filename = f"../logs/excel/{init_time_formatted}-log.xlsx"

        if os.path.exists(self.filename):
            self.workbook = load_workbook(self.filename)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            headers = ["Gold Value", "Mineral Value", "Is Worth Attacking", "Uptime", "Looted Gold", "Looted Minerals",
                       "Gold efficiency", "Mineral efficiency"]
            for col, header in enumerate(headers, start=1):
                self.sheet.cell(row=1, column=col, value=header)

        self.workbook.save(self.filename)

    def log_to_excel(self, gold_value, mineral_value, is_worth, uptime, loot_gold_value, loot_mineral_value):
        """
        Logs data to the Excel spreadsheet.

        Params:
            gold_value (str): The gold value
            mineral_value (str): The mineral value
            is_worth (bool): True if the base is worth attacking, False otherwise
            uptime (timedelta): The bot's uptime
            loot_gold_value (str): Amount of looted gold
            loot_mineral_value (str): Amount of looted minerals
        """
        new_row = [gold_value, mineral_value, is_worth, uptime, loot_gold_value, loot_mineral_value]
        self.sheet.append(new_row)
        self.workbook.save(self.filename)


# TODO: list below
"""
# TODO: calculate efficiency?
as for efficiencies, i want them to be stored in new columns within our log table
mineral_efficiency = loot_mineral_value / mineral_value
gold_efficiency = loot_gold_value / gold_value
total_efficiency = (loot_mineral_value+loot_gold_value)/(mineral_value+gold_value)
at some point (maybe when starting the logger, we should save values assigned to threshholds for:
defensive_buildings, gold and mineral value (like the one that we use to decide if attack is worth it)
i want it to be printed over and over again for each entry in a separate column
-------------------
i want to add some kind of total logging that will be used to calculate stats for entire project
i want to keep logging stuff for each run separately, but i also want them to be logged to a one bigger logging
file (we will append new entries there)
like yk log1.xlsx log2.xlsx log3.xlsx for each run and all_logs.xlsx for file containing all the runs together
remember about new columns (especially the efficiency ones) and the fact that we have naming convention already
-------------------
i want all the logging to console be logged to some file, let's say logs.txt; new logs are gonna be appended to the file
####gotta fix it -> logs are cleared every time
"""
