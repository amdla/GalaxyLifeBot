import logging
import os
import time
from datetime import datetime

import cv2
import easyocr
import numpy as np
import pyautogui
import torch
from openpyxl import Workbook, load_workbook
from pygetwindow import getWindowsWithTitle
from ultralytics import YOLO

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Define button coordinates as constants
ATTACK_BUTTON = (935, 1365)
FIND_TARGET_BUTTON = (950, 1270)
FIGHT_NOW_BUTTON = (1360, 865)
SEARCH_AGAIN_BUTTON = (1030, 1305)
TRAINING_CAMP_1_BUTTON = (1265, 505)
TRAINING_CAMP_2_BUTTON = (1390, 565)
CLOSE_TRAINING_VIEW_BUTTON = (1615, 1340)
SPEED_UP_X2_BUTTON = (1535, 160)
END_BATTLE_BUTTON = (1475, 200)
GO_HOME_BUTTON = (1250, 935)
OPEN_PLANETS_LIST_BUTTON = (940, 1320)
COLONY_11_BUTTON = (1190, 1240)
ADD_LOOTERS_TO_TRAINING_LIST_BUTTON = (1070, 1420)
CHOOSE_LOOTER_UNIT_WHEN_ATTACKING_BUTTON = (950, 1375)
CLOSE_NEWS_POPUP_BUTTON = (1583, 204)
CLOSE_DAILY_GIFT_POPUP_BUTTON = (1590, 530)

global init_time
global workbook, sheet


def init_excel_logging():
    """
    Initializes the Excel workbook for logging.
    """
    global workbook, sheet

    filename = "bot_logs.xlsx"
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bot Logs"
        headers = ["Timestamp", "Gold Value", "Mineral Value", "Is Worth Attacking", "Uptime"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    workbook.save(filename)


def log_to_excel(gold_value, mineral_value, is_worth, uptime):
    """
    Logs data to the Excel spreadsheet.
    """
    global workbook, sheet

    new_row = [gold_value, mineral_value, is_worth, str(uptime)]
    sheet.append(new_row)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{timestamp}-log.xlsx"
    workbook.save(file_name)


def click_and_wait(button, time_to_wait):
    """
    Clicks a specified button and waits for a given amount of time.

    Parameters:
    button (tuple): The (x, y) coordinates of the button to click
    time_to_wait (float): The number of seconds to wait after clicking
    """

    pyautogui.click(button)
    time.sleep(time_to_wait)


def handle_error():
    """
    Handles errors by refreshing the game window and resetting to the initial base.
    """

    get_screenshot("Galaxy Life")  # used just to take focus on window
    logging.warning("---------------------------------Handling error with F5 refresh---------------------------------")
    pyautogui.keyDown('F5')
    time.sleep(0.2)
    pyautogui.keyUp('F5')
    time.sleep(8)
    click_and_wait(CLOSE_NEWS_POPUP_BUTTON, 8)
    click_and_wait(CLOSE_DAILY_GIFT_POPUP_BUTTON, 2)

    get_initial_base()


def get_initial_base():
    """
    Navigates to the initial base by opening the planets list and selecting Colony 11.
    """
    try:
        click_and_wait(OPEN_PLANETS_LIST_BUTTON, 2)
        click_and_wait(COLONY_11_BUTTON, 12)
    except Exception as e:
        raise e


def parse_detection_file(result_path):
    """
    Parses a detection file to extract base coordinates and detections.

    Parameters:
    result_path (str): The path to the detection results file

    Returns:
    tuple(base_coords, detections) where:
        base_coords (tuple): Coordinates of the base or None if they are not found #TODO tuple of 4 floats
        detections (list): List of detection coordinates #TODO list of 4x floats
    """

    try:
        with open(result_path, 'r') as file:
            lines = file.readlines()

        base_coords = None
        detections = []

        for line in lines:
            parts = line.strip().split(',')
            if len(parts) == 6:
                class_id, score, x1, y1, x2, y2 = map(float, parts)
                if int(class_id) == 7:
                    base_coords = (x1, y1, x2, y2)
                detections.append((x1, y1, x2, y2))

        return base_coords, detections
    except Exception as e:
        raise e


def calculate_detections_deltas(detections, base_coords):
    """
    Calculates deltas of detections, and determines if the base is on the edge.

    Parameters:
    detections (list): List of detection coordinates #TODO list of 4x floats
    base_coords (tuple): Coordinates of the base or None if they are not found #TODO tuple of 4 floats

    Returns:
    tuple(deltas, is_base_on_edge) where:
        deltas (tuple): The delta values for the rectangle corners #TODO tuple of 4 floats
        is_base_on_edge (bool): True if the base is on the edge
    """
    # TODO: decide whether delta should be for all detections or just these with score > 0.33
    try:
        if not detections or base_coords is None:
            return

        x1_values = [coords[0] for coords in detections]
        x2_values = [coords[2] for coords in detections]
        y1_values = [coords[1] for coords in detections]
        y2_values = [coords[3] for coords in detections]

        deltas = [
            min(x1_values),
            max(x2_values),
            min(y1_values),
            max(y2_values)
        ]

        is_base_on_edge = (base_coords and
                           (base_coords[0] == deltas[0] or base_coords[2] == deltas[1] or
                            base_coords[1] == deltas[2] or base_coords[3] == deltas[3]))

        return deltas, is_base_on_edge
    except Exception as e:
        raise e


def save_detection_results(screen_path, results):
    """
    Saves detection results to a file and draws bounding boxes on the image.

    Parameters:
    screen_path (str): Path to the screenshot image
    results: Detection results from the YOLO model

    Returns:
    tuple (result_path, image) where:
        result_path(str): Path to the detection results file
        image: Image with bounding boxes drawn
    """
    try:
        result_path = f"{screen_path}.txt"
        image = cv2.imread(screen_path)

        with open(result_path, 'w') as f:
            f.write(f"Total detected boxes: {len(results.boxes.data)}\n")
            f.write("Detected boxes and their scores:\n")
            for result in results.boxes.data.tolist():
                x1, y1, x2, y2, score, class_id = result
                f.write(f"{class_id},{score},{x1},{y1},{x2},{y2}\n")
                if score > 0.33:
                    cv2.rectangle(image, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 4)
                    cv2.putText(image, results.names[int(class_id)].upper(), (int(x1), int(y1) - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 255, 0), 2, cv2.LINE_AA)

        return result_path, image
    except Exception as e:
        raise e


def draw_encompassing_rectangle(image, deltas):
    """
    Draws an encompassing rectangle on the image based on delta values.

    Parameters:
    image: The image to draw on
    deltas (tuple): The delta values for the rectangle corners #TODO tuple of 4 floats
    """
    try:
        delta_x1, delta_x2, delta_y1, delta_y2 = deltas
        cv2.rectangle(image, (int(delta_x1), int(delta_y1)), (int(delta_x2), int(delta_y2)), (255, 0, 0), 4)
    except Exception as e:
        raise e


def is_worth_based_on_defences(screen_path):
    """
    Determines if a base is worth attacking based on defences recognized by model.

    Parameters:
    screen_path (str): Path to the screenshot image

    Returns:
    bool: True if the base is worth attacking, False otherwise
    """
    try:
        image = cv2.imread(screen_path)
        if image is None:
            raise FileNotFoundError(f"Image not found at {screen_path}")
        model_path = "../runs/detect/train104/weights/best.pt"
        model = YOLO(model_path)
        results = model(image)[0]
        result_path, image = save_detection_results(screen_path, results)
        base_coords, detections = parse_detection_file(result_path)
        deltas, is_base_on_edge = calculate_detections_deltas(detections, base_coords)
        if deltas:
            draw_encompassing_rectangle(image, tuple(deltas))
            cv2.imwrite(f"{screen_path}_detections.png", image)
        else:
            logging.error("No valid boxes found to calculate average location and deltas.")
            return False
        # worth attacking when base is on edge or there are less than 5 defensive buildings
        result = len(results.boxes.data) < 4 or is_base_on_edge
        logging.info(f"Is worth attacking based on defences: {result}")
        return result
    except Exception as e:
        raise e


def is_worth_attacking(values_list):
    """
    Determines if a base is worth attacking based on resources and defences.

    Parameters:
    values_list (list): A list containing gold value, mineral value, and screen path #TODO: float, float, str

    Returns:
    bool: True if the base is worth attacking, False otherwise
    """
    try:
        gold_value, mineral_value, screen_path = values_list
        result = int(mineral_value) > 800000 and is_worth_based_on_defences(screen_path)
        logging.info(f"Is worth attacking based on resources: {int(mineral_value) > 800000}")
        logging.info(f"Is worth attacking: {result}")
        logging.info("---------------------------ended calculating, proceeds to next enemy---------------------------")

        sheet.cell(row=sheet.max_row, column=4, value=str(result))
        workbook.save("bot_logs.xlsx")

        return result
    except Exception as e:
        raise e


def search_for_enemy():
    """
    Performs the sequence of clicks to search for an enemy base.
    """
    try:
        click_and_wait(ATTACK_BUTTON, 2.5)
        click_and_wait(FIND_TARGET_BUTTON, 2.5)
        click_and_wait(FIGHT_NOW_BUTTON, 2.5)
        click_and_wait(FIGHT_NOW_BUTTON, 8)
    except Exception as e:
        raise e


def extract_region_of_interest(image, x, y, width, height):
    """
    Extracts a region of interest from an image.

    Parameters:
    image: The source image
    x (int): X-coordinate of the top-left corner
    y (int): Y-coordinate of the top-left corner
    width (int): Width of the region
    height (int): Height of the region

    Returns:
    numpy.ndarray: The extracted region of interest
    """
    try:
        return image[y:y + height, x:x + width]
    except Exception as e:
        raise e


def split_region_of_interest_for_gold_and_mineral_fields(region_of_interest):
    """
    Splits a region of interest into two parts for gold and mineral fields.

    Parameters:
    region_of_interest (numpy.ndarray): The region of interest to split

    Returns:
    tuple: Two numpy arrays representing the split regions
    """
    try:
        mid_index = region_of_interest.shape[0] // 2
        return region_of_interest[:mid_index, :], region_of_interest[mid_index:, :]
    except Exception as e:
        raise e


def read_resource_values(region_of_interest, reader):
    """
    Reads resource values from a region of interest using OCR.

    Parameters:
    region_of_interest (numpy.ndarray): The region to read from
    reader: An EasyOCR reader object

    Returns:
    str: The extracted numeric value as a string
    """
    try:
        result = reader.readtext(np.array(region_of_interest), detail=0)
        return ''.join(filter(str.isdigit, ''.join(result)))
    except Exception as e:
        raise e


def get_gold_and_minerals(screenshot):
    """
    Extracts gold and mineral values from a screenshot.

    Parameters:
    screenshot: The screenshot image

    Returns:
    tuple(gold_value, mineral_value) where:
        gold_value (str): The extracted gold value
        mineral_value (str): The extracted mineral value
    """
    try:
        screenshot_np = np.array(screenshot.convert('L'))
        region_of_interest = extract_region_of_interest(screenshot_np, 985, 100, 65, 50)

        with torch.no_grad():
            reader = easyocr.Reader(['en'], gpu=torch.cuda.is_available())

        split_regions_of_interest = split_region_of_interest_for_gold_and_mineral_fields(region_of_interest)
        gold_value = read_resource_values(split_regions_of_interest[0], reader)
        mineral_value = read_resource_values(split_regions_of_interest[1], reader)

        return gold_value, mineral_value
    except Exception as e:
        raise e


def get_screenshot(window_title):
    """
    Captures a screenshot of a specified window.

    Parameters:
    window_title (str): The title of the window to capture

    Returns:
    PIL.Image: The captured screenshot
    """
    try:
        window = getWindowsWithTitle(window_title)[0]
        window.activate()
        time.sleep(0.25)
        window.maximize()
        time.sleep(0.25)
        window.moveTo(0, 0)
        time.sleep(0.25)
    except IndexError:
        logging.error("Window not found!")
        exit()
    return pyautogui.screenshot()


def process_screenshot():
    """
    Processes a screenshot to extract gold and mineral values. Saves bot's uptime to a file.

    Returns:
    list: A list containing gold value, mineral value, and the path to the saved screenshot #TODO: float, float, str
    """
    try:
        window_title = "Galaxy Life"
        screenshot = get_screenshot(window_title)

        if screenshot:
            now = datetime.now()
            uptime = now - init_time
            logging.info(f"Uptime: {uptime}")

            timestamp = now.strftime("%Y%m%d_%H%M%S")
            screen_path = f"../../screenshots/screen{timestamp}.png"
            screenshot.save(screen_path)

            gold_value, mineral_value = get_gold_and_minerals(screenshot)

            logging.info(f"Gold Value: {gold_value}")
            logging.info(f"Mineral Value: {mineral_value}")

            log_to_excel(gold_value, mineral_value, "", str(uptime))

            return [gold_value, mineral_value, screen_path]

    except Exception as e:
        raise e


def deploy_troops():
    """
    Deploys troops in a specific pattern around the enemy base.
    """
    try:
        click_and_wait(CHOOSE_LOOTER_UNIT_WHEN_ATTACKING_BUTTON, 0.5)
        coordinates = [
            (1460, 290), (1695, 435), (1870, 525), (1980, 590), (2090, 620),
            (2210, 685), (2340, 800), (2210, 870), (2060, 955), (1915, 1030),
            (1800, 1090), (1650, 1150), (1420, 1230), (1135, 1220), (855, 1160),
            (600, 1055), (490, 930), (445, 740), (480, 620), (630, 500),
            (840, 405), (1000, 310)
        ]
        for x, y in coordinates:
            click_and_wait((x, y), 0.1)
    except Exception as e:
        raise e


def add_troops_to_training():
    """
    Adds troops to the training queue in the training camps.
    """
    try:
        click_and_wait(TRAINING_CAMP_1_BUTTON, 0.1)
        click_and_wait(TRAINING_CAMP_1_BUTTON, 1.5)
        for _ in range(15):
            click_and_wait(ADD_LOOTERS_TO_TRAINING_LIST_BUTTON, 0.1)
        click_and_wait(TRAINING_CAMP_2_BUTTON, 0.1)
        click_and_wait(TRAINING_CAMP_2_BUTTON, 1.5)
        for _ in range(15):
            click_and_wait(ADD_LOOTERS_TO_TRAINING_LIST_BUTTON, 0.1)
        click_and_wait(CLOSE_TRAINING_VIEW_BUTTON, 1)
    except Exception as e:
        raise e


def attack():
    """
    Performs an attack sequence including troop deployment and battle management.
    """
    try:
        deploy_troops()
        click_and_wait(SPEED_UP_X2_BUTTON, 0)
        click_and_wait(SPEED_UP_X2_BUTTON, 50)  # Wait for the battle to end (x4 speed)
        click_and_wait(END_BATTLE_BUTTON, 5)
        click_and_wait(GO_HOME_BUTTON, 10)

    except Exception as e:
        raise e


def main_loop():
    """
    The main game loop that manages the overall flow of searching for enemies and attacking.
    """
    iterations = 0
    init_excel_logging()
    while True:
        try:
            search_for_enemy()
            while True:
                iterations += 1
                if iterations % 50 == 0:
                    raise Exception("Restarting the game after 50 iterations to avoid getting stuck.")
                values_list = process_screenshot()
                if is_worth_attacking(values_list):
                    attack()
                    add_troops_to_training()
                    break
                else:
                    click_and_wait(SEARCH_AGAIN_BUTTON, 10)
        except Exception as e:
            logging.error(f"Error: {e}")
            iterations = 0
            handle_error()


if __name__ == '__main__':
    get_screenshot("Galaxy Life")  # used just to take focus on window
    init_time = datetime.now()
    main_loop()
